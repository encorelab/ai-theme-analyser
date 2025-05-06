import os
import re
import json
from collections import defaultdict
import matplotlib.pyplot as plt
import networkx as nx
import docx
import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import time
import logging
import math
import networkx as nx
import datetime
import ast
from config import *
import tiktoken


# Configure logging
LOG_FILE = "log.txt"
logging.basicConfig(filename=LOG_FILE, level=logging.INFO, 
                    format="%(asctime)s - %(levelname)s - %(message)s")


def remove_json_markdown(text):
    """Removes JSON markdown from a string."""
    pattern = re.compile(r'```json\s*(.*?)\s*```', re.DOTALL)
    return pattern.sub(r'\1', text)


def extract_paragraphs_from_docx(filepath):
    """
    Extracts paragraphs from a docx file, formats them with markdown bolding for headings,
    and returns them as a list of strings.
    """
    doc = docx.Document(filepath)
    formatted_paragraphs = []
    for p in doc.paragraphs:
        if p.text.strip():
            if p.style.name.startswith('Heading 1') or p.style.name.startswith('Heading 2'):
                formatted_paragraphs.append(f"**{p.text}**")
            else:
                formatted_paragraphs.append(p.text)
    return formatted_paragraphs


def write_coding_results_to_excel(all_files_excerpt_codings: dict[str, dict[str, list[str]]], 
                                 new_codes_by_file: dict[str, dict[str, dict]], 
                                 output_file: str):
    """
    Writes the coding results and justifications to an Excel file.
    Ensures the Excel file is created even if the data is blank.
    """

    print(f"Received all_files_excerpt_codings: {all_files_excerpt_codings}")
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Prepare data for Excel output (codings) for ALL files
            all_codings_data = []
            for filename, excerpt_codings in all_files_excerpt_codings.items():
                for excerpt, codes in excerpt_codings.items():
                    # Remove illegal characters from excerpt
                    cleaned_excerpt = ILLEGAL_CHARACTERS_RE.sub(r'', excerpt)  
                    all_codings_data.append({
                        'filename': filename,
                        'excerpt': cleaned_excerpt,  # Use the cleaned excerpt
                        'codings': ', '.join(codes)
                    })

            # Write codings to Excel for ALL files (even if empty)
            codings_df = pd.DataFrame(all_codings_data)  # Create DataFrame even if empty
            codings_df.to_excel(writer, sheet_name='codings', index=False)
            writer.sheets['codings'].sheet_state = 'visible' 

            # Prepare data for Excel output (new_codes)
            all_justifications_data = []
            print("Exporting new codes by file to code_justifications sheet:\n" + json.dumps(new_codes_by_file, indent=4)) 
            for filename, new_codes in new_codes_by_file.items():
                for code, data in new_codes.items():

                    all_justifications_data.append({
                        'code': code,
                        'filename': filename,
                        'examples': data['excerpt'],
                        'construct': data.get('theme', ''),
                        'description': data.get('description', ''),
                        'justification': data['justification'],
                        'probability': data['probability']
                    })

            # Write code justifications to Excel (even if empty)
            justifications_df = pd.DataFrame(all_justifications_data)  # Create DataFrame even if empty
            justifications_df.to_excel(writer, sheet_name='code_justifications', index=False)
            writer.sheets['code_justifications'].sheet_state = 'visible' 

    except Exception as e:
        print(f"An error occurred while writing to Excel: {e}")


def load_themes_from_file(filepath):
    """
    Loads themes and their definitions from a JSON file.
    """
    try:
        with open(filepath, 'r') as f:
            themes_data = json.load(f)
        
        themes = []
        theme_definitions = {}
        for item in themes_data:
            themes.append(item["theme"])
            theme_definitions[item["theme"]] = item["definition"]

        return themes, theme_definitions

    except FileNotFoundError:
        print(f"Error: File not found - {filepath}")
        return [], {}
    
    
def generate_codes(directory,
                    themes,
                    coding_client,
                    initial_codes=None,
                    words_per_chunk=1200,
                    num_docs=None,
                    time_between_calls=6):
    """
    Generates initial codes with dynamic delay to avoid rate limiting.

    Args:
        directory: Directory with docx files.
        themes: Dictionary of code constructs and definitions.
        coding_client: CodeGeneratorClient instance.
        initial_themes: starting set of codes, if any.
        words_per_chunk: Approximate words per chunk.
        num_docs: Number of documents to process (all if None).
        time_between_calls: Target time between calls in seconds.

    Returns:
        Tuple of coding results.
    """
    start_time = time.time()

    if initial_codes is None:
        all_codes = {}
    else:
        all_codes = initial_codes
    
    all_files_excerpt_codings = {}
    new_codes_by_file = {}

    docx_files = [
        f for f in os.listdir(directory)
        if f.endswith('.docx') and not f.startswith('~$')
    ]
    if num_docs is not None:
        docx_files = docx_files[:num_docs]
    total_files = len(docx_files)
    processed_files = 0

    for filename in docx_files:
        filepath = os.path.join(directory, filename)
        if filename.endswith('.docx') and not filename.startswith('~$'):
            paragraphs = extract_paragraphs_from_docx(filepath)
            file_excerpt_codings = {}
            paragraph_chunks = chunk_paragraphs(paragraphs, words_per_chunk)

            for chunk in paragraph_chunks:
                for construct in themes:
                    call_start_time = time.time()  # Record call start time

                    excerpt_codings, all_codes, new_codes = generate_codes_for_chunk(
                        chunk, construct, coding_client, all_codes
                    )

                    # Merge excerpt_codings into file_excerpt_codings
                    for excerpt, codes in excerpt_codings.items():
                        if excerpt in file_excerpt_codings:
                            for code in codes:
                                if code not in file_excerpt_codings[excerpt]:
                                    file_excerpt_codings[excerpt].append(code)
                        else:
                            file_excerpt_codings[excerpt] = codes

                    # Append new_codes to the existing codes for the filename
                    if filename not in new_codes_by_file:
                        new_codes_by_file[filename] = {}
                    new_codes_by_file[filename].update(new_codes)

                    call_end_time = time.time()  # Record call end time
                    call_duration = call_end_time - call_start_time

                    # Calculate dynamic delay
                    delay = max(0, time_between_calls - call_duration) 
                    time.sleep(delay)

            all_files_excerpt_codings[filename] = file_excerpt_codings

        # --- Print timestamp and progress ---
        processed_files += 1
        elapsed_time = time.time() - start_time
        remaining_files = total_files - processed_files
        print(
            f"Processed {filename} ({processed_files}/{total_files} files). Time elapsed: {elapsed_time:.2f} seconds. Remaining: {remaining_files} files."
        )
    return all_codes, all_files_excerpt_codings, new_codes_by_file

def chunk_paragraphs(paragraphs, words_per_chunk=1200):
    """Chunks a list of paragraphs into smaller lists based on word count."""
    paragraph_chunks = []
    current_chunk = []
    current_word_count = 0
    for paragraph in paragraphs:
        current_chunk.append(paragraph)
        current_word_count += len(paragraph.split())
        if current_word_count >= words_per_chunk:
            paragraph_chunks.append(current_chunk)
            current_chunk = []
            current_word_count = 0
    if current_chunk:  # Add the last chunk if not empty
        paragraph_chunks.append(current_chunk)
    return paragraph_chunks


def generate_codes_for_chunk(chunk, construct, code_generation_client, all_codes):
    """Generates codes for a single chunk of text and accumulates the results."""
    # Create data_for_chunk as a string with markdown formatting
    data_for_chunk = ""
    for paragraph in chunk:
        data_for_chunk += f"{paragraph}\n\n"

    # Call generate_initial_codes with construct definition and formatted string
    excerpt_codings, new_codes = code_generation_client.generate_codes(
                        data_for_chunk, all_codes, construct 
                    )

    # Add new codes and descriptions to the ongoing lists
    for code, data in new_codes.items():
        if code not in all_codes:
            all_codes[code] = data.copy()

    return excerpt_codings, all_codes, new_codes


def perform_analysis_and_reporting(output_file, themes, analyzer_client, intra_text_analyzer):
    """
    Performs thematic and intra-text analysis on coded data and writes results to Excel.
    """
    codings_df = pd.read_excel(output_file, sheet_name='codings')

    analysis_output_file = os.path.join(OUTPUT_DIR, f"thematic_analysis_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
    intra_text_output_file = os.path.join(OUTPUT_DIR, f"intra_text_analysis_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")

    with pd.ExcelWriter(analysis_output_file) as analysis_writer, pd.ExcelWriter(intra_text_output_file) as intra_text_writer:
        for filename in codings_df['filename'].unique():
            file_data = codings_df[codings_df['filename'] == filename]
            excerpts = {
                filename: {
                    row['paragraph_index']: row['original_text']
                    for _, row in file_data.iterrows()
                }
            }

            analysis_results = analyzer_client.analyze_themes(json.dumps(excerpts), themes)
            themes = analysis_results['themes_to_codes']

            # Write thematic analysis results to Excel
            analysis_df = pd.DataFrame([
                {
                    'filename': filename,
                    'paragraph_index': paragraph_index,
                    'themes': ', '.join(analysis_data['themes']),
                    'quote': analysis_data['quote'],
                    'justification': analysis_data['justification']
                }
                for paragraph_index, analysis_data in analysis_results['analysis'].items()
            ])
            analysis_df.to_excel(analysis_writer, sheet_name='analysis', index=False)

            # Perform intra-text analysis
            intra_text_results = intra_text_analyzer.analyze_intra_text(json.dumps(analysis_results))

            # Write intra-text analysis results to Excel
            for analysis_type in ['intersections', 'contradictions', 'connections']:
                intra_text_df = pd.DataFrame(intra_text_results[analysis_type])
                if not intra_text_df.empty:
                    intra_text_df.insert(0, 'filename', filename)
                    intra_text_df.insert(1, 'analysis_type', analysis_type)
                    intra_text_df.to_excel(intra_text_writer, sheet_name='intra_text', index=False, header=False, startrow=intra_text_writer.sheets['intra_text'].max_row)

    return intra_text_output_file


def perform_cross_document_analysis(intra_text_output_file, cross_document_analyzer):
    """
    Performs cross-document analysis on intra-text results and prints the summary.
    """
    intra_text_df = pd.read_excel(intra_text_output_file, sheet_name='intra_text')
    cross_document_summary = cross_document_analyzer.analyze_cross_document(intra_text_df.to_json(orient='records'))

    print("\nCross-Document Analysis Summary:\n")
    print(cross_document_summary)


def perform_intra_text_analysis(analysis_results, intra_text_analyzer):
    """
    Performs intra-text analysis on the provided analysis results and writes 
    the output to an Excel file.
    """
    intra_text_output_file = os.path.join(OUTPUT_DIR, f"intra_text_analysis_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")

    with pd.ExcelWriter(intra_text_output_file) as intra_text_writer:
        for filename, analysis_data in analysis_results.items():
            intra_text_results = intra_text_analyzer.analyze_intra_text(json.dumps(analysis_data))

            # Write intra-text analysis results to Excel
            for analysis_type in ['intersections', 'contradictions', 'connections']:
                intra_text_df = pd.DataFrame(intra_text_results[analysis_type])
                if not intra_text_df.empty:
                    intra_text_df.insert(0, 'filename', filename)
                    intra_text_df.insert(1, 'analysis_type', analysis_type)
                    intra_text_df.to_excel(intra_text_writer, sheet_name='intra_text', index=False, header=False, startrow=intra_text_writer.sheets['intra_text'].max_row)

    return intra_text_output_file


def load_analysis_results_from_file(filepath):
    """
    Loads thematic analysis results from a JSON file.
    """
    try:
        with open(filepath, 'r') as f:
            analysis_results = json.load(f)
        return analysis_results
    except FileNotFoundError:
        print(f"Error: File not found - {filepath}")
        return {}


def load_themes_from_file(filepath):
    """
    Loads themes from a JSON file.
    """
    try:
        with open(filepath, 'r') as f:
            themes = json.load(f)
        return themes
    except FileNotFoundError:
        print(f"Error: File not found - {filepath}")
        return {}


def load_codes_from_file(filepath):
    """
    Loads codes and their descriptions and constructs from a JSON file.
    """
    try:
        with open(filepath, 'r') as f:
            code_data = json.load(f)

        return code_data

    except FileNotFoundError:
        print(f"Error: File not found - {filepath}")
        return {}
    
def load_codes_from_file_as_dictionary(filepath):
    """
    Loads codes and their descriptions and constructs from a JSON file.
    Transforms the list of dictionaries into a dictionary where the code name is the key.
    """
    try:
        with open(filepath, 'r') as f:
            code_data = json.load(f)

        transformed_codes = {}
        for code_entry in code_data:
            code_name = code_entry["code"]
            transformed_codes[code_name] = {
                "description": code_entry["description"],
                "theme": code_entry["construct"],
                "examples": code_entry.get("examples", ""),  # Handle cases where examples might be missing
                "exclude": code_entry.get("exclude", "")  # Handle cases where exclude might be missing
            }

        return transformed_codes

    except FileNotFoundError:
        print(f"Error: File not found - {filepath}")
        return {}
    

def convert_codes_dict_to_dataframe(codes_dict):
    """
    Converts a dictionary of codes into a Pandas DataFrame.
    Handles 'examples' as either a string or a list of strings.
    """
    data = []
    for code_name, code_details in codes_dict.items():
        examples = code_details.get("examples", [])

        # Handle 'examples' as either a string or a list
        if isinstance(examples, str):
            examples = [examples]  # Treat single string example as a list with one element

        if not examples:
            examples = [""]  # Handle cases with no examples

        for example in examples:
            data.append({
                "code": code_name,
                "excerpt": example,  # Each example becomes a separate row
                "theme": code_details.get("theme", ""),
                "description": code_details.get("description", ""),
                "justification": code_details.get("justification", "") 
            })

    return pd.DataFrame(data)


def visualize_theme_overview(themes_hierarchy, filename="class1_theme_overview.png"):
    """
    Visualizes the overview of meta-themes, themes, and sub-themes and saves it as an image.
    Node sizes are adjusted based on their frequency.

    Args:
        themes_hierarchy: The hierarchical theme structure.
        filename: The name of the file to save the visualization.
    """

    graph = nx.DiGraph()
    node_labels = {}
    node_colors = {}
    color_map = plt.get_cmap("tab20")

    # --- Scaling parameters (adjust these as needed) ---
    scaling_factor = 1000  # Controls how much the frequency affects the size
    base_size = 200     # Minimum size of a node

    def get_node_size(frequency):
        """Calculates node size based on frequency using a logarithmic scale."""
        return base_size + scaling_factor * math.log(frequency + 1)

    def add_nodes_and_edges(hierarchy, level=0, parent=None):
        """Recursively adds nodes and edges to the graph."""
        for name, data in hierarchy.items():
            node_id = str(name)

            # Get frequency and calculate node size
            frequency = data.get("frequency", 1)
            node_size = get_node_size(frequency)

            graph.add_node(node_id, size=node_size) # Store size as node attribute
            node_labels[node_id] = name
            node_colors[node_id] = color_map(level)

            if parent:
                graph.add_edge(parent, node_id)

            # Recursively add children (themes or sub-themes)
            if "themes" in data:
                add_nodes_and_edges(data["themes"], level + 1, node_id)
            if "sub-themes" in data:
                add_nodes_and_edges(data["sub-themes"], level + 1, node_id)

    # Build the graph
    add_nodes_and_edges(themes_hierarchy)

    plt.figure(figsize=(24, 12))
    pos = nx.spring_layout(graph, k=0.3)

    # Draw the graph, using the calculated sizes
    nx.draw(graph,
            pos,
            labels=node_labels,
            with_labels=True,
            node_size=[d['size'] for n, d in graph.nodes(data=True)], # Get sizes from node attributes
            node_color=[node_colors[node] for node in graph.nodes()],
            font_size=10,
            font_weight="bold",
            arrowsize=20)

    plt.savefig(filename)
    # plt.show()


def visualize_individual_theme_subgraphs(themes_hierarchy, output_dir="theme_subgraphs"):
    """
    Visualizes subgraphs for each theme within the themes hierarchy,
    adjusting node sizes based on frequency, and saves each visualization
    as a separate image file.

    Args:
        themes_hierarchy: The hierarchical theme structure.
        output_dir: The directory to save the visualization files.
    """

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # --- Scaling parameters (adjust these as needed) ---
    scaling_factor = 2000  # Controls how much the frequency affects the size
    base_size = 200  # Minimum size of a node

    def get_node_size(frequency):
        """Calculates node size based on frequency using a logarithmic scale."""
        return base_size + scaling_factor * math.log(frequency + 1)

    # Iterate through each meta-theme and then each theme to create individual subgraphs
    for meta_theme, meta_theme_data in themes_hierarchy.items():
        for theme, theme_data in meta_theme_data.get("themes", {}).items():
            graph = nx.DiGraph()
            node_labels = {}
            node_colors = {}
            color_map = plt.get_cmap("tab20")

            def add_nodes_and_edges(parent, data, level=0):
                """Recursively adds nodes and edges to the graph."""
                for name, sub_data in data.items():
                    node_id = str(name)

                    # Get frequency of sub-theme and calculate node size
                    sub_theme_frequency = sub_data.get("frequency", 1)
                    node_size = get_node_size(sub_theme_frequency)

                    graph.add_node(node_id, size=node_size)
                    node_labels[node_id] = name
                    node_colors[node_id] = color_map(level + 1)

                    if parent:
                        graph.add_edge(parent, node_id)

                    # Recursively add children (sub-themes or codes)
                    if "sub-themes" in sub_data:
                        add_nodes_and_edges(node_id, sub_data["sub-themes"], level + 1)
                    if "codes" in sub_data:
                        # Use code frequencies for node sizes
                        code_frequencies = sub_data.get("code_frequencies", {})
                        for code, code_frequency in code_frequencies.items():
                            code_id = str(code)
                            code_size = get_node_size(code_frequency)
                            graph.add_node(code_id, size=code_size)
                            node_labels[code_id] = code.split("-", 1)[-1]
                            node_colors[code_id] = color_map(level + 2)
                            graph.add_edge(node_id, code_id)

            # Add the current theme's data to the graph
            add_nodes_and_edges(None, {theme: theme_data})

            # Generate filename based on theme name and timestamp
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{theme}_{timestamp}.png"
            filepath = os.path.join(output_dir, filename)

            # Draw and save the subgraph
            plt.figure(figsize=(24, 8))
            pos = nx.spring_layout(graph, k=0.5)
            nx.draw(graph,
                    pos,
                    labels=node_labels,
                    with_labels=True,
                    node_size=[d['size'] for n, d in graph.nodes(data=True)],
                    node_color=[node_colors[node] for node in graph.nodes()],
                    font_size=15,
                    font_weight="bold",
                    arrowsize=20)
            plt.title(f"Subgraph for Theme: {theme}\n(Meta-Theme: {meta_theme})")
            plt.margins(x=0.15) # Adds 15% padding on right/left sides
            plt.savefig(filepath)
            plt.close()

            print(f"Saved subgraph for theme '{theme}' to '{filepath}'")


def visualize_single_file_graph(
    filtered_themes_hierarchy, filename_to_analyze, output_dir="within_case_network_graphs"
):
    """
    Visualizes a single network graph for a specific file, including all relevant
    themes, sub-themes, and codes, with node sizes adjusted based on frequency.

    Args:
        filtered_themes_hierarchy: The filtered theme hierarchy containing only the relevant data for the file.
        filename_to_analyze: The name of the file being analyzed (e.g., "104.docx").
        output_dir: The directory to save the visualization file.
    """

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # --- Scaling parameters (adjust these as needed) ---
    scaling_factor = 1000  # Controls how much the frequency affects the size
    base_size = 200  # Minimum size of a node

    def get_node_size(frequency):
        """Calculates node size based on frequency using a logarithmic scale."""
        return base_size + scaling_factor * math.log(frequency + 1)

    graph = nx.DiGraph()
    node_labels = {}
    node_colors = {}
    color_map = plt.get_cmap("tab20")

    def add_nodes_and_edges(parent, data, level=0):
        """Recursively adds nodes and edges to the graph."""
        for name, sub_data in data.items():
            node_id = str(name)

            # Get frequency and calculate node size
            frequency = sub_data.get("frequency", 1)
            node_size = get_node_size(frequency)

            graph.add_node(node_id, size=node_size)
            node_labels[node_id] = name
            node_colors[node_id] = color_map(level)

            if parent:
                graph.add_edge(parent, node_id)

            # Recursively add children (themes, sub-themes, or codes)
            if "themes" in sub_data:
                add_nodes_and_edges(node_id, sub_data["themes"], level + 1)
            if "sub-themes" in sub_data:
                add_nodes_and_edges(node_id, sub_data["sub-themes"], level + 1)
            if "codes" in sub_data:
                # Use code frequencies for node sizes (if available)
                code_frequencies = sub_data.get("code_frequencies", {})
                for code in sub_data["codes"]:
                    code_id = str(code)
                    code_frequency = code_frequencies.get(code, 1)
                    code_size = get_node_size(code_frequency)
                    graph.add_node(code_id, size=code_size)
                    node_labels[code_id] = code
                    node_colors[code_id] = color_map(level + 1)
                    graph.add_edge(node_id, code_id)

    # Add all relevant data to the graph
    add_nodes_and_edges(None, filtered_themes_hierarchy)

    # Generate filename based on analyzed filename and timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"{filename_to_analyze}_{timestamp}.png"
    filepath = os.path.join(output_dir, output_filename)

    # Draw and save the graph
    plt.figure(figsize=(24, 12))  # Increase figure size for better visibility
    pos = nx.spring_layout(graph, k=0.3)
    nx.draw(
        graph,
        pos,
        labels=node_labels,
        with_labels=True,
        node_size=[d["size"] for n, d in graph.nodes(data=True)],
        node_color=[node_colors[node] for node in graph.nodes()],
        font_size=8,  # Adjust font size if needed
        font_weight="bold",
        arrowsize=15,
    )
    plt.title(f"Network Graph for File: {filename_to_analyze}")
    plt.savefig(filepath)
    plt.close()

    print(f"Saved network graph for '{filename_to_analyze}' to '{filepath}'")

def visualize_network(data, filename="network_visualization.png"):
    """
    Visualizes a network graph from a JSON object containing nodes and edges.

    Args:
        data: A JSON object with "nodes" and "edges" lists.
        filename: The name of the file to save the visualization.
    """

    graph = nx.DiGraph()
    node_labels = {}
    node_colors = {}
    edge_labels = {}
    color_map = plt.get_cmap("tab20")

    # Add nodes to the graph
    for i, node in enumerate(data["nodes"]):
        node_id = node["id"]
        graph.add_node(node_id)
        node_labels[node_id] = node["label"]
        node_colors[node_id] = color_map(i)  # Assign color based on index

    # Add edges to the graph
    for edge in data["edges"]:
        source = edge["source"]
        target = edge["target"]
        relation = edge["relation"]
        graph.add_edge(source, target)
        edge_labels[(source, target)] = relation

    # Set figure size and layout
    plt.figure(figsize=(24, 8))
    pos = nx.spring_layout(graph, k=0.3)

    # Draw nodes with labels and colors
    nx.draw(graph,
            pos,
            labels=node_labels,
            with_labels=True,
            node_size=3000,
            node_color=[node_colors[node] for node in graph.nodes()],
            font_size=10,
            font_weight="bold",
            arrowsize=20)

    # Draw edge labels
    nx.draw_networkx_edge_labels(graph, pos, edge_labels=edge_labels, font_size=8)

    plt.savefig(filename)
    plt.show()

def read_full_dataset_codes(file_path):
  """
  Reads an xlsx file, extracts data from the first two sheets, and returns them as separate pandas DataFrames.

  Args:
    file_path: The path to the xlsx file.

  Returns:
    A tuple containing two pandas DataFrames:
      - all_codings: DataFrame from the first sheet (named 'codings') with columns 'filename', 'excerpt', and 'codings'.
      - new_codes: DataFrame from the second sheet (named 'code_justifications') with columns 'code', 'filename', 'excerpt', 'theme', 'description', 'justification', and 'probability'.
  """
  try:
    # Read the xlsx file, specifying that we want to read only the first two sheets.
    # We use sheet_name=None to read all sheets and then select the first two.
    xlsx = pd.read_excel(file_path, sheet_name=None)

    # Extract the first sheet as 'all_codings'
    all_codings = xlsx[list(xlsx.keys())[0]]  
    all_codings.columns = ['filename', 'excerpt', 'codings']

    # Extract the second sheet as 'new_codes'
    new_codes = xlsx[list(xlsx.keys())[1]]
    new_codes.columns = ['code', 'filename', 'excerpt', 'theme', 'description', 'justification', 'probability']

    return all_codings, new_codes

  except FileNotFoundError:
    print(f"Error: File not found at path: {file_path}")
    return None, None
  except ValueError as e:
      if "Worksheet" in str(e):
          print(f"Error: The specified xlsx file does not have at least two sheets.")
      else:
        print(f"Error: Invalid file format or other issue with reading the xlsx file: {e}")
      return None, None
  except Exception as e:
    print(f"An unexpected error occurred: {e}")
    return None, None


def extract_unique_used_codes(codings_df):
    """
    Extracts unique codes from the 'codings' column of a DataFrame.

    Args:
        codings_df: DataFrame with a 'codings' column containing comma-separated codes.

    Returns:
        DataFrame with a 'unique_used_code_name' column containing unique codes.
    """
    unique_codes = set()
    for index, row in codings_df.iterrows():
        codings = row['codings']
        if pd.notna(codings):  # Check if the value is not NaN
            codes = str(codings).split(',')  # Convert to string to handle different data types
            for code in codes:
                unique_codes.add(code.strip())

    unique_codes_df = pd.DataFrame({'unique_used_code_name': sorted(list(unique_codes))})
    return unique_codes_df

def generate_code_stats(full_dataset_file_path, initial_codes_file_path, output_filepath):
    """
    Generates code statistics, creates a new Excel workbook with multiple sheets,
    and prints duplicate codes to the console.

    Args:
        full_dataset_file_path: Path to the full dataset Excel file.
        initial_codes_file_path: Path to the JSON file containing initial codes.
        output_filepath: Path to save the output Excel file.
    """
    try:
        # 1. Read data and prepare DataFrames
        all_codings, new_codes = read_full_dataset_codes(full_dataset_file_path)
        initial_codes_dict = load_codes_from_file_as_dictionary(initial_codes_file_path)
        initial_codes_df = convert_codes_dict_to_dataframe(initial_codes_dict)

        # 2. Create 'used_codes' DataFrame
        used_codes_df = extract_unique_used_codes(all_codings)
        used_codes_df['is_initial_code'] = used_codes_df['unique_used_code_name'].isin(initial_codes_df['code'])
        used_codes_df['is_new_code'] = used_codes_df['unique_used_code_name'].isin(new_codes['code'])
        used_codes_df['is_defined_code'] = used_codes_df['is_initial_code'] | used_codes_df['is_new_code']

        # 3. Calculate statistics (before creating used_codes_with_def)
        total_unique_codes_used = len(used_codes_df)
        total_initial_codes = len(initial_codes_df)
        total_initial_codes_used = used_codes_df['is_initial_code'].sum()
        total_new_codes = len(new_codes['code'].unique())
        total_undefined_codes = (~used_codes_df['is_defined_code']).sum()

        stats_df = pd.DataFrame({
            'Statistic': ['Total Unique Codes Used', 'Total Initial Codes', 'Total Initial Codes Used',
                          'Total New Codes', 'Total Undefined Codes'],
            'Value': [total_unique_codes_used, total_initial_codes, total_initial_codes_used,
                      total_new_codes, total_undefined_codes]
        })

        # 4. Prepare initial_codes and new_codes for merging
        initial_codes_df = initial_codes_df.rename(
            columns={"excerpt": "examples", "theme": "construct"}
        )
        if "justification" in initial_codes_df.columns: # Remove justification if it is there.
            initial_codes_df = initial_codes_df.drop(columns=["justification"])
        new_codes = new_codes.rename(
            columns={"excerpt": "examples", "theme": "construct"}
        )

        # --- Combine initial_codes and new_codes ---
        combined_codes_df = pd.concat([initial_codes_df, new_codes], ignore_index=True)
        # Drop duplicate rows based on 'code', keeping the first occurrence (likely from initial_codes)
        combined_codes_df = combined_codes_df.drop_duplicates(subset=['code'], keep='first')


        # 5. Create used_codes_with_def_df by merging
        used_codes_with_def_df = used_codes_df.rename(columns={"unique_used_code_name": "code"})
        used_codes_with_def_df = used_codes_with_def_df.merge(
            combined_codes_df[["code", "description", "examples", "construct"]],
            on="code",
            how="left",  # Important: LEFT JOIN to keep all used_codes
        )

        # Calculate frequency (do this *after* the merge)
        code_frequency = {}
        for index, row in all_codings.iterrows():
            codings = row['codings']
            if pd.notna(codings):
                codes = str(codings).split(',')
                for code in codes:
                    code = code.strip()
                    code_frequency[code] = code_frequency.get(code, 0) + 1
        used_codes_with_def_df['frequency'] = used_codes_with_def_df['code'].map(code_frequency).fillna(0)


        # --- Check for undefined codes ---
        for index, row in used_codes_with_def_df.iterrows():
            if pd.isna(row["description"]) and pd.isna(row["examples"]) and pd.isna(row["construct"]):
                print(f"Error: Code '{row['code']}' is used but not defined in initial_codes or new_codes.")


        # 6. Write to Excel
        with pd.ExcelWriter(output_filepath, engine="openpyxl") as writer:
            used_codes_with_def_df.to_excel(
                writer, sheet_name="used_codes_with_def", index=False
            )
            all_codings.to_excel(writer, sheet_name="codings", index=False)
            initial_codes_df.to_excel(writer, sheet_name="initial_codes", index=False)
            new_codes.to_excel(writer, sheet_name="new_codes", index=False)
            used_codes_df.to_excel(writer, sheet_name="used_codes", index=False)
            stats_df.to_excel(writer, sheet_name="stats", index=False)

        print(
            f"Successfully generated code statistics and saved to '{output_filepath}'"
        )

    except FileNotFoundError:
        print(f"Error: File not found.")
    except ValueError as e:
        print(f"Error: Invalid file format - {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

def read_used_codes_with_def(file_path):
    """Reads the 'used_codes_with_def' sheet from an Excel file."""
    try:
        workbook = pd.read_excel(file_path, sheet_name="used_codes_with_def")
        # Check for required columns
        required_columns = ["code", "description", "examples", "construct"]
        if not all(col in workbook.columns for col in required_columns):
            raise ValueError(
                "The 'used_codes_with_def' sheet must contain columns: 'code', 'description', 'examples', 'construct'"
            )
        return workbook
    except FileNotFoundError:
        print(f"Error: File not found at path: {file_path}")
        return None
    except ValueError as e:
        print(f"Error: {e}")  # More specific error message
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None


def convert_df_to_codes_dict(df):
    """Converts the DataFrame to a dictionary with codes as keys."""
    codes_dict = {}
    for _, row in df.iterrows():
        code = row["code"]
        codes_dict[code] = {
            "description": row["description"],
            "theme": row["construct"],
            "examples": row["examples"],
        }
    return codes_dict


def replace_and_update_codes(
    full_dataset_file_path, merged_codes_file_path, output_filepath
):
    """
    Replaces merged codes in the full dataset codings and updates the
    used_codes_with_def sheet to reflect the merged codes.

    Args:
        full_dataset_file_path: Path to the Excel file with full dataset.
        merged_codes_file_path: Path to the Excel file with merged codes.
        output_filepath: Path to save the updated Excel file.
    """
    try:
        # --- 1. Read Data ---
        # Read full dataset codings
        try:
            full_dataset_df = pd.read_excel(
                full_dataset_file_path, sheet_name="codings"
            )
        except ValueError as e:
            if "Worksheet named 'codings' not found" in str(e):
                print(
                    f"Error: File '{full_dataset_file_path}' lacks 'codings' sheet."
                )
                return
            else:
                raise

        # Read merged codes
        try:
            merged_codes_df = pd.read_excel(
                merged_codes_file_path, sheet_name="Merged Codes"
            )
        except ValueError as e:
            if "Worksheet named 'Merged Codes' not found" in str(e):
                print(
                    f"Error: File '{merged_codes_file_path}' lacks 'Merged Codes' sheet."
                )
                return
            else:
                raise

        # Read used_codes_with_def (from full_dataset file)
        try:
            used_codes_df = pd.read_excel(
                full_dataset_file_path, sheet_name="used_codes_with_def"
            )
        except ValueError as e:
            if "Worksheet named 'used_codes_with_def' not found" in str(e):
                print(
                    "Error: 'used_codes_with_def' sheet not found in full dataset file."
                )
                return
            else:
                raise

        # --- 2. Create Code Mapping ---
        code_map = {}
        for _, row in merged_codes_df.iterrows():
            new_code = row["code"]
            merged_codes_str = row["merged_codes"]

            if isinstance(merged_codes_str, str):
                try:
                    merged_codes_list = ast.literal_eval(merged_codes_str)
                except (ValueError, SyntaxError):
                    print(
                        f"Warning: Invalid list format in 'merged_codes' at index {row.name}."
                    )
                    continue
            elif isinstance(merged_codes_str, list):
                merged_codes_list = merged_codes_str
            else:
                print(
                    f"Warning: Unexpected type in 'merged_codes': {type(merged_codes_str)} at index {row.name}."
                )
                continue

            for old_code in merged_codes_list:
                code_map[old_code.strip()] = new_code.strip()

        # --- 3. Replace Codes in Full Dataset ---
        def replace_codes(codings_str):
            if pd.isna(codings_str):
                return ""
            codings_str = str(codings_str)
            new_codes = []
            for code in codings_str.split(","):
                code = code.strip()
                new_code = code_map.get(code, code)  # Replace or keep original
                new_codes.append(new_code)
            return ", ".join(new_codes)

        full_dataset_df["codings"] = full_dataset_df["codings"].apply(replace_codes)

        # --- 4. Update used_codes_with_def ---

        # 4.a. Remove rows with old codes
        merged_codes_set = set()
        for _, row in merged_codes_df.iterrows():
             if isinstance(row["merged_codes"], str):
                try:
                    merged_codes_list = ast.literal_eval(row["merged_codes"])
                except (ValueError, SyntaxError):
                    print(f"Warning: Invalid list format in 'merged_codes' at index {row.name}. Skipping this row.")
                    continue  # Skip to the next row
             elif isinstance(row["merged_codes"], list):
                 merged_codes_list = row["merged_codes"]
             for code in merged_codes_list:
                merged_codes_set.add(code)
        
        used_codes_updated_df = used_codes_df[
            ~used_codes_df["code"].isin(merged_codes_set)
        ]

        # 4.b. Add new rows for merged codes
        new_rows = []
        for _, row in merged_codes_df.iterrows():
            new_rows.append(
                {
                    "code": row["code"],
                    "description": row["description"],
                    "examples": row["examples"],
                    "construct": "",  # Assuming construct is not in merged_codes
                    "frequency": 0,  # Placeholder, will be recalculated
                }
            )
        used_codes_updated_df = pd.concat(
            [used_codes_updated_df, pd.DataFrame(new_rows)], ignore_index=True
        )

        # 4.c Recalculate Frequencies (Important!)
        code_frequency = {}
        for index, row in full_dataset_df.iterrows(): #Use the UPDATED full dataset
            codings = row["codings"]
            if pd.notna(codings):
                codes = str(codings).split(",")
                for code in codes:
                    code = code.strip()
                    code_frequency[code] = code_frequency.get(code, 0) + 1
        used_codes_updated_df["frequency"] = used_codes_updated_df["code"].map(
            code_frequency
        ).fillna(0)


        # --- 5. Write to Excel ---
        with pd.ExcelWriter(output_filepath, engine="openpyxl") as writer:
            full_dataset_df.to_excel(writer, sheet_name="Merged Codings", index=False)
            used_codes_updated_df.to_excel(
                writer, sheet_name="Updated Used Codes", index=False
            )  

        print(f"Successfully processed and saved to '{output_filepath}'")

    except FileNotFoundError as e:
        print(f"Error: File not found - {e.filename}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def split_data_by_class(merged_codings_file_path):
    """
    Splits data into separate files based on LCA class membership.

    Args:
        merged_codings_file_path: Path to the Excel file with merged codings
                                   and class information.
    """
    try:
        # --- 1. Read Data ---
        try:
            merged_codings_df = pd.read_excel(
                merged_codings_file_path, sheet_name="Merged Codings"
            )
        except ValueError as e:
            if "Worksheet named 'Merged Codings' not found" in str(e):
                print(
                    f"Error: File '{merged_codings_file_path}' lacks 'Merged Codings' sheet."
                )
                return
            else:
                raise

        try:
            updated_used_codes_df = pd.read_excel(
                merged_codings_file_path, sheet_name="Updated Used Codes"
            )
        except ValueError as e:
            if "Worksheet named 'Updated Used Codes' not found" in str(e):
                print(
                    "Error: 'Updated Used Codes' sheet not found in merged codings file."
                )
                return
            else:
                raise

        # --- 2. Calculate Codes per Construct (Original) ---
        original_codes_per_construct = (
            updated_used_codes_df.groupby("construct")["code"]
            .nunique()
            .reset_index(name="num_codes")
        )
        original_codes_per_construct.rename(
            columns={"construct": "unique_construct"}, inplace=True
        )

        # --- 3. Get Unique Classes ---
        if "class" not in merged_codings_df.columns:
            print("Error: 'class' column not found in 'Merged Codings' sheet.")
            return
        unique_classes = merged_codings_df["class"].unique()

        # --- 4. Split and Save Data ---
        for class_val in unique_classes:
            # 4.a. Filter Merged Codings (using .copy())
            filtered_merged_codings = merged_codings_df[
                merged_codings_df["class"] == class_val
            ].copy()

            # 4.b. Filter Updated Used Codes (based on filtered Merged Codings, using .copy())
            relevant_codes = set()
            for codings_str in filtered_merged_codings["codings"]:
                if pd.notna(codings_str):
                    codes = str(codings_str).split(",")
                    for code in codes:
                        relevant_codes.add(code.strip())

            filtered_used_codes = updated_used_codes_df[
                updated_used_codes_df["code"].isin(relevant_codes)
            ].copy()

            # 4.c. Recalculate Frequencies
            code_frequency = {}
            for _, row in filtered_merged_codings.iterrows():
                codings = row["codings"]
                if pd.notna(codings):
                    codes = str(codings).split(",")
                    for code in codes:
                        code = code.strip()
                        code_frequency[code] = code_frequency.get(code, 0) + 1

            filtered_used_codes["frequency"] = filtered_used_codes["code"].map(
                code_frequency
            ).fillna(0)

            # 4.d. Calculate Codes per Construct (Filtered)
            filtered_codes_per_construct = (
                filtered_used_codes.groupby("construct")["code"]
                .nunique()
                .reset_index(name="num_codes")
            )
            filtered_codes_per_construct.rename(
                columns={"construct": "unique_construct"}, inplace=True
            )

            # 4.e. Save to File
            output_filename = f"class_{class_val}_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            output_filepath = os.path.join(OUTPUT_DIR, output_filename)

            with pd.ExcelWriter(output_filepath, engine="openpyxl") as writer:
                filtered_merged_codings.to_excel(
                    writer, sheet_name="Merged Codings", index=False
                )
                filtered_used_codes.to_excel(
                    writer, sheet_name="Updated Used Codes", index=False
                )
                filtered_codes_per_construct.to_excel(
                    writer, sheet_name="Codes per Construct", index=False
                )

            print(f"Successfully created class file: '{output_filepath}'")

        # --- 5. Add original codes per construct sheet ---
        try:
            with pd.ExcelWriter(
                merged_codings_file_path, engine="openpyxl", mode="a"
            ) as writer:
                if "Codes per Construct" in writer.book.sheetnames:
                    del writer.book["Codes per Construct"]
                original_codes_per_construct.to_excel(
                    writer, sheet_name="Codes per Construct", index=False
                )
            print(
                "Successfully created sheet 'Codes per Construct' in file provided by filepath."
            )

        except Exception as e:
            print(f"An error occurred while creating a sheet: {e}")

    except FileNotFoundError as e:
        print(f"Error: File not found - {e.filename}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def count_tokens(text: str) -> int:
    """Counts the number of tokens in a string using the cl100k_base encoding."""
    try:
        encoding = tiktoken.get_encoding("cl100k_base")  # Or other appropriate encoding
        num_tokens = len(encoding.encode(text))
        return num_tokens
    except Exception as e:
        print(f"Error in token counting: {e}")
        return -1 # Return -1 to indicate an error.


def split_codes_list(codes_list):
    """Splits a list of code dictionaries into two approximately equal halves."""
    midpoint = len(codes_list) // 2
    first_half = codes_list[:midpoint]
    second_half = codes_list[midpoint:]
    return first_half, second_half

def compress_code_examples(codes_file_path, compression_type, compressor_client):
    """Compresses code examples in a JSON file using an LLM."""
    MAX_TOKEN_SIZE = LARGE_GENERATION_CONFIG['max_output_tokens'] - 1000
    try:
        codes_dict = load_codes_from_file_as_list_of_dict(codes_file_path)
        if not codes_dict:
            print("Error: Could not load codes from file.")
            return
        
        if compression_type not in ("1", "2"):
            print("Invalid compression type. Please enter 1 or 2.")
            return

        # Initial token count
        initial_tokens = count_tokens(json.dumps(codes_dict))
        print(f"Initial token count: {initial_tokens}")

        # Split and process until all chunks are small enough
        chunks_to_process = [codes_dict]
        processed_chunks = []

        while chunks_to_process:
            current_chunk = chunks_to_process.pop(0)
            chunk_tokens = count_tokens(json.dumps(current_chunk))

            if chunk_tokens > MAX_TOKEN_SIZE:
                print(f"Splitting chunk (size: {chunk_tokens})")
                first_half, second_half = split_codes_list(current_chunk)
                chunks_to_process.extend([first_half, second_half])
            else:
                print(f"Processing chunk (size: {chunk_tokens})")
                processed_chunks.append(current_chunk)

        compressed_results = []  # Accumulate results as a LIST

        # Process each small-enough chunk
        for chunk in processed_chunks:
            compressed_chunk = compressor_client.compress_examples(chunk, compression_type)

            # Check for valid JSON response before merging
            if isinstance(compressed_chunk, list):  # Expect a LIST now
                compressed_results.extend(compressed_chunk)  # Extend the list
            else:
                print(f"Warning: Invalid JSON response from compressor: {compressed_chunk}")
                # Handle the error appropriately, e.g., skip, retry, or log

        # Final token count
        final_tokens = count_tokens(json.dumps(compressed_results))
        print(f"Final token count: {final_tokens}")

        if final_tokens < 32768:
            print("Successfully compressed code examples.")
        else:
            print("Warning: Final JSON is still larger than 32768 tokens.")

        # Construct output file path
        base_name, ext = os.path.splitext(codes_file_path)
        output_filepath = f"{base_name}_compressed{ext}"

        # Save the compressed JSON
        with open(output_filepath, 'w', encoding='utf-8') as f:
            json.dump(compressed_results, f, indent=4) # Dump the LIST

        print(f"Compressed codes saved to: {output_filepath}")

    except FileNotFoundError:
        print(f"Error: File not found - {codes_file_path}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def load_codes_from_file_as_list_of_dict(filepath):
    """
    Loads codes and their descriptions and constructs from a JSON file.
    Now returns the original list format, not a dictionary keyed by code.
    """
    try:
        with open(filepath, 'r') as f:
            code_data = json.load(f)
        return code_data  # Return the list directly

    except FileNotFoundError:
        print(f"Error: File not found - {filepath}")
        return []  # Return empty list on error
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON format in {filepath}")
        return []